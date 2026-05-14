"""RAG Evaluation — оценка качества системы по тестовым вопросам.

Запуск:
    python tests/rag_eval/evaluator.py
    python tests/rag_eval/evaluator.py --no-ragas --no-ragchecker
    python tests/rag_eval/evaluator.py --api-url http://192.168.1.10
"""

import argparse
import asyncio
import json
import logging
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import httpx
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from tests.rag_eval.metrics.bertscore_metric import compute_bertscore
from tests.rag_eval.metrics.ragas_metrics import compute_ragas
from tests.rag_eval.metrics.ragchecker_metric import compute_ragchecker

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

REPORTS_DIR = Path(__file__).parent / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

CATEGORY_ALIASES: dict[str, str] = {
    "sql": "SQL", "sqL": "SQL", "SQl": "SQL", "SQL": "SQL",
    "vector": "Vector", "Vector": "Vector",
    "compomat": "Compomat", "Compomat": "Compomat",
    "compromat": "Compromat", "Compromat": "Compromat",
}


def normalize_category(raw: str) -> str:
    return CATEGORY_ALIASES.get(raw.strip(), raw.strip())


def load_questions(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Файл пустой: {path}")

    logger.info("Колонки: %s", rows[0])
    questions = []
    for row in rows[1:]:
        if not row[0]:
            continue
        questions.append({
            "question": str(row[0]).strip(),
            "category": normalize_category(str(row[1]).strip() if row[1] else "Unknown"),
            "ground_truth": str(row[2]).strip() if row[2] else "",
        })
    logger.info("Загружено %d вопросов", len(questions))
    return questions


async def ask_api(client: httpx.AsyncClient, api_url: str, question: str, user_id: str) -> dict:
    try:
        resp = await client.post(
            f"{api_url}/v1/chat",
            json={"user_id": user_id, "question": question},
            timeout=120,
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as exc:
        logger.error("API error: %s", exc)
        return {"answer": "", "source": "error", "error": str(exc)}


async def aggregate_metrics(
    results: list[dict],
    use_ragas: bool,
    use_ragchecker: bool,
    llm_url: str = "http://localhost:9002/v1",
) -> dict:
    successful = [r for r in results if r.get("actual_answer") and not r.get("error")]

    metrics: dict = {
        "count": len(results),
        "successful": len(successful),
        "error_rate": round((len(results) - len(successful)) / len(results), 4) if results else 0,
    }

    if not successful:
        metrics["error"] = "no successful results"
        return metrics

    predictions = [r["actual_answer"] for r in successful]
    references = [r["ground_truth"] for r in successful]

    metrics["bertscore"] = await compute_bertscore(predictions, references)

    if use_ragas:
        metrics["ragas"] = await compute_ragas(successful, llm_url=llm_url)

    if use_ragchecker:
        metrics["ragchecker"] = await compute_ragchecker(successful, llm_url=llm_url)

    return metrics


def print_metrics(label: str, metrics: dict) -> None:
    print(f"\n{'=' * 55}")
    print(f"  {label}")
    print(f"{'=' * 55}")

    if "error" in metrics and not metrics.get("count"):
        print(f"  ⚠ {metrics['error']}")
        return

    print(f"  Вопросов:       {metrics['count']}")
    print(f"  Успешных:       {metrics['successful']}")
    print(f"  Error rate:     {metrics['error_rate']:.1%}")

    bs = metrics.get("bertscore", {})
    if bs and "error" not in bs:
        print(f"  Cosine F1:      {bs.get('f1_mean', 0):.4f}  "
              f"(min={bs.get('f1_min', 0):.4f}, max={bs.get('f1_max', 0):.4f})")

    ragas = metrics.get("ragas", {})
    if ragas and "error" not in ragas:
        print(f"  Faithfulness:   {ragas.get('faithfulness', 0):.4f}")
        print(f"  Answer relev.:  {ragas.get('answer_relevancy', 0):.4f}")
        print(f"  Context recall: {ragas.get('context_recall', 0):.4f}")

    rc = metrics.get("ragchecker", {})
    if rc and "error" not in rc:
        print(f"  Claim recall:   {rc.get('claim_recall', 0):.4f}")
        print(f"  Claim prec.:    {rc.get('claim_precision', 0):.4f}")
        print(f"  Hallucination:  {rc.get('hallucination_rate', 0):.4f}")


async def evaluate(
    questions_path: Path,
    api_url: str,
    use_ragas: bool,
    use_ragchecker: bool,
    llm_url: str = "http://localhost:9002/v1",
) -> dict:
    questions = load_questions(questions_path)
    categories = sorted({q["category"] for q in questions})
    print(f"\nЗапускаем оценку: {len(questions)} вопросов → {api_url}/v1/chat")
    print(f"Категории: {categories}\n")

    results = []
    async with httpx.AsyncClient() as client:
        for i, q in enumerate(questions, 1):
            print(
                f"[{i:2d}/{len(questions)}] [{q['category']:10s}] {q['question'][:55]}...",
                end=" ", flush=True,
            )
            await asyncio.sleep(2)
            response = await ask_api(client, api_url, q["question"], f"eval_{i}")
            actual = response.get("answer", "")
            result = {
                "question": q["question"],
                "category": q["category"],
                "ground_truth": q["ground_truth"],
                "actual_answer": actual,
                "contexts": [f["text"] for f in response.get("fragments", [])],
                "source": response.get("source", "unknown"),
                "error": response.get("error"),
            }
            results.append(result)
            print("✓" if actual and not result.get("error") else "✗")

    print("\nВычисляем метрики...")
    overall = await aggregate_metrics(results, use_ragas, use_ragchecker, llm_url)

    by_category: dict[str, list] = defaultdict(list)
    for r in results:
        by_category[r["category"]].append(r)

    category_metrics = {}
    for cat, cat_results in sorted(by_category.items()):
        category_metrics[cat] = await aggregate_metrics(cat_results, use_ragas, use_ragchecker, llm_url)

    print_metrics("ОБЩИЕ МЕТРИКИ", overall)
    for cat, m in category_metrics.items():
        print_metrics(f"Категория: {cat}", m)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = REPORTS_DIR / f"eval_{timestamp}.json"
    report = {
        "timestamp": datetime.now().isoformat(),
        "questions_file": str(questions_path),
        "api_url": api_url,
        "overall": overall,
        "by_category": category_metrics,
        "results": results,
    }
    with report_path.open("w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print(f"\n✓ Отчёт сохранён: {report_path}")
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="RAG Evaluation")
    parser.add_argument("--questions", type=Path,
                        default=Path("tests/rag_eval/data/questions.xlsx"))
    parser.add_argument("--api-url", default="http://localhost")
    parser.add_argument("--no-ragas", action="store_true")
    parser.add_argument("--no-ragchecker", action="store_true")
    parser.add_argument("--llm-url", default="http://localhost:9002/v1",
                        help="URL локального vLLM для RAGAS/RAGChecker")
    args = parser.parse_args()

    if not args.questions.exists():
        print(f"Error: файл не найден: {args.questions}")
        sys.exit(1)

    asyncio.run(evaluate(
        args.questions,
        args.api_url,
        not args.no_ragas,
        not args.no_ragchecker,
        args.llm_url,
    ))


if __name__ == "__main__":
    main()