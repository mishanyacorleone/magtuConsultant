"""
Скрипт 01: Сбор ответов от RAG-системы.

Категории берутся напрямую из Excel без нормализации — только lower().

Вход:  tests/rag_eval/data/questions.xlsx
Выход: tests/rag_eval/results/raw_answers.json
"""

import argparse
import asyncio
import json
import sys
from datetime import datetime
from pathlib import Path

import httpx
from openpyxl import load_workbook

RESULTS_DIR = Path(__file__).parent.parent / "results"
DATA_DIR = Path(__file__).parent.parent / "data"


def load_questions(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    questions = []
    for i, row in enumerate(rows[1:], start=1):
        if not row[0]:
            continue
        questions.append({
            "query_id": f"q_{i:03d}",
            "question": str(row[0]).strip(),
            # Категория — как есть из Excel, только lower()
            "category": str(row[1]).strip().lower() if row[1] else "unknown",
            "ground_truth": str(row[2]).strip() if row[2] else "",
        })
    return questions


async def ask(client: httpx.AsyncClient, api_url: str, question: str, user_id: str) -> dict:
    try:
        resp = await client.post(
            f"{api_url}/v1/chat",
            json={"user_id": user_id, "question": question},
            timeout=120,
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as exc:
        return {"answer": "", "source": "error", "fragments": [], "error": str(exc)}


async def collect(questions_path: Path, api_url: str, delay: float) -> list[dict]:
    questions = load_questions(questions_path)
    categories = sorted({q["category"] for q in questions})
    print(f"Загружено {len(questions)} вопросов")
    print(f"Категории: {categories}")
    print(f"API: {api_url}  |  Задержка: {delay}s\n")

    results = []
    async with httpx.AsyncClient() as client:
        for i, q in enumerate(questions, 1):
            print(f"[{i:2d}/{len(questions)}] [{q['category']:12s}] {q['question'][:55]}...",
                  end=" ", flush=True)

            if delay > 0:
                await asyncio.sleep(delay)

            response = await ask(client, api_url, q["question"], q["query_id"])

            retrieved_context = [
                {"text": f["text"], "source": f.get("source", ""), "score": f.get("score", 0)}
                for f in response.get("fragments", [])
            ]

            record = {
                "query_id": q["query_id"],
                "question": q["question"],
                "category": q["category"],
                "ground_truth": q["ground_truth"],
                "actual_answer": response.get("answer", ""),
                "source": response.get("source", "unknown"),
                "retrieved_context": retrieved_context,
                "error": response.get("error"),
            }
            results.append(record)
            print("✓" if record["actual_answer"] and not record.get("error") else "✗")

    return results


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--questions", type=Path, default=DATA_DIR / "questions.xlsx")
    parser.add_argument("--api-url", default="http://localhost")
    parser.add_argument("--delay", type=float, default=2.0)
    parser.add_argument("--output", type=Path, default=RESULTS_DIR / "raw_answers.json")
    args = parser.parse_args()

    if not args.questions.exists():
        print(f"Ошибка: {args.questions} не найден")
        sys.exit(1)

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    results = asyncio.run(collect(args.questions, args.api_url, args.delay))

    successful = sum(1 for r in results if r["actual_answer"] and not r.get("error"))
    print(f"\nУспешно: {successful}/{len(results)}")

    output = {
        "collected_at": datetime.now().isoformat(),
        "api_url": args.api_url,
        "total": len(results),
        "successful": successful,
        "results": results,
    }

    with args.output.open("w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"✓ Сохранено: {args.output}")


if __name__ == "__main__":
    main()