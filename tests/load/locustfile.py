"""
Нагрузочное тестирование RAG-системы через Locust.

Вопросы берутся из questions.xlsx — каждый запрос использует случайный вопрос.
Категория логируется для анализа производительности по типам.

Установка:
    pip install locust openpyxl

Запуск (web UI):
    locust -f tests/load/locustfile.py --host http://localhost

Запуск без UI (headless):
    locust -f tests/load/locustfile.py --host http://localhost \
        --headless --users 10 --spawn-rate 2 --run-time 60s

Параметры:
    --users N       — количество одновременных пользователей
    --spawn-rate N  — сколько новых пользователей добавлять в секунду
    --run-time Xs   — длительность теста
"""

import json
import random
from pathlib import Path

from locust import HttpUser, between, task, events
from openpyxl import load_workbook

QUESTIONS_PATH = Path(__file__).parent.parent / "rag_eval" / "data" / "questions.xlsx"


def _load_questions() -> list[dict]:
    """Загружает вопросы из Excel при старте."""
    if not QUESTIONS_PATH.exists():
        print(f"[WARN] Файл вопросов не найден: {QUESTIONS_PATH}")
        return [
            {"question": "Какие специальности есть в МГТУ?", "category": "fallback"},
            {"question": "Какой проходной балл на информатику?", "category": "fallback"},
        ]

    wb = load_workbook(QUESTIONS_PATH, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    questions = []
    for row in rows[1:]:
        if row[0]:
            questions.append({
                "question": str(row[0]).strip(),
                "category": str(row[1]).strip().lower() if row[1] else "unknown",
            })
    print(f"[INFO] Загружено {len(questions)} вопросов для нагрузочного теста")
    return questions


# Загружаем вопросы один раз при импорте модуля
QUESTIONS = _load_questions()

# Счётчики по категориям для статистики
_category_stats: dict[str, dict] = {}


@events.request.add_listener
def on_request(request_type, name, response_time, response_length,
               exception, context, **kwargs):
    """Собирает статистику по категориям."""
    cat = context.get("category", "unknown") if context else "unknown"
    if cat not in _category_stats:
        _category_stats[cat] = {"count": 0, "failures": 0, "total_ms": 0}
    _category_stats[cat]["count"] += 1
    _category_stats[cat]["total_ms"] += response_time
    if exception:
        _category_stats[cat]["failures"] += 1


@events.test_stop.add_listener
def on_test_stop(environment, **kwargs):
    """Выводит статистику по категориям после теста."""
    if not _category_stats:
        return
    print("\n" + "=" * 60)
    print("  СТАТИСТИКА ПО КАТЕГОРИЯМ")
    print("=" * 60)
    for cat, stats in sorted(_category_stats.items()):
        count = stats["count"]
        failures = stats["failures"]
        avg_ms = stats["total_ms"] / count if count > 0 else 0
        print(f"  {cat:<15} запросов={count:4d}  ошибок={failures:3d}"
              f"  avg={avg_ms:7.0f}ms")
    print("=" * 60)


class RAGUser(HttpUser):
    """Симулирует пользователя который задаёт случайные вопросы."""

    wait_time = between(1, 5)  # пауза между запросами одного пользователя

    def on_start(self):
        """Каждый виртуальный пользователь получает уникальный ID."""
        self.user_id = f"load_test_{random.randint(100000, 999999)}"

    @task
    def ask_question(self):
        """Отправляет случайный вопрос из Excel файла."""
        q = random.choice(QUESTIONS)

        payload = {
            "user_id": self.user_id,
            "question": q["question"],
        }

        with self.client.post(
            "/v1/chat",
            json=payload,
            catch_response=True,
            context={"category": q["category"]},
            name=f"/v1/chat [{q['category']}]",
        ) as response:
            if response.status_code == 200:
                try:
                    data = response.json()
                    if not data.get("answer"):
                        response.failure("Пустой ответ")
                    else:
                        response.success()
                except Exception as e:
                    response.failure(f"Ошибка парсинга JSON: {e}")
            elif response.status_code == 503:
                response.failure("Circuit breaker открыт (503)")
            else:
                response.failure(f"HTTP {response.status_code}")